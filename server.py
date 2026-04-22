from __future__ import annotations

import base64
import io
import json
import hashlib
import os
import re
import subprocess
import threading
import urllib.parse
import urllib.request
import uuid
from datetime import datetime
from pathlib import Path

from flask import Flask, abort, jsonify, request, send_from_directory, send_file
from flask_compress import Compress
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent
WEB_DIR = ROOT / "web"
HOSPITAL_LIST_PATH = ROOT / "블로그목록.md"
CONFIG_PATH = ROOT / "config" / "april_keywords.json"
ZEROPAIN_CONFIG_PATH = ROOT / "config" / "zeropain_april_keywords.json"
SAMSUNGBON_CONFIG_PATH = ROOT / "config" / "samsungbon_march_keywords.json"
JL_CONFIG_PATH = ROOT / "config" / "jl_march_keywords.json"
SNU_CONFIG_PATH = ROOT / "config" / "snu_march_keywords.json"
SCRIPT_PATH = ROOT / "scripts" / "build_april_month.py"
DATA_PATH = ROOT / "data" / "scoring-data.json"

# build_april_month.SHEETS_META 와 동일 순서·키 (런타임 시트 배치용)
SHEET_ORDER: list[tuple[str, str]] = [
    ("regional-pc", "2026 지역 PC"),
    ("regional-mob", "2026 지역 MOB"),
    ("national-pc", "2026 전국 PC"),
    ("national-mob", "2026 전국 MOB"),
    ("other-pc", "2026 기타 PC"),
    ("other-mob", "2026 기타 MOB"),
]

# 블로그목록 식별명과 scoring-data 의 hospitalName 이 다를 때 (예: SNU서울정형외과 → SNU서울병원)
HOSPITAL_CANONICAL: dict[str, str] = {
    "SNU서울정형외과": "SNU서울병원",
}

# 배포 안정성: 채점은 비동기 작업으로 분리해서 API 응답 타임아웃/중복 실행을 방지.
SCORE_TASKS: dict[str, dict[str, object]] = {}
ACTIVE_SCORE_TASK_BY_KEY: dict[str, str] = {}
SCORE_TASKS_LOCK = threading.Lock()
MERGE_LOCK = threading.Lock()  # scoring-data.json / last-run-evidence.json 동시 쓰기 방지


def canonical_hospital_name(name: str | None) -> str:
    n = (name or "").strip() or "포인트병원"
    return HOSPITAL_CANONICAL.get(n, n)


def template_config_path(hospital_name: str) -> Path:
    """병원별 기본 채점 설정 파일(키워드 없을 때 런타임 구성의 베이스)."""
    raw = (hospital_name or "").strip()
    name = canonical_hospital_name(raw)
    if raw == "제로마취통증의학과" and ZEROPAIN_CONFIG_PATH.exists():
        return ZEROPAIN_CONFIG_PATH
    if raw in {"삼성본병원", "삼성본정형외과"} and SAMSUNGBON_CONFIG_PATH.exists():
        return SAMSUNGBON_CONFIG_PATH
    if raw == "제이엘정형외과" and JL_CONFIG_PATH.exists():
        return JL_CONFIG_PATH
    if name == "SNU서울병원" and SNU_CONFIG_PATH.exists():
        return SNU_CONFIG_PATH
    return CONFIG_PATH

HOSPITAL_PROFILE_OVERRIDES: dict[str, dict[str, object]] = {
    "포인트병원": {},
    "클래스병원": {
        "hospitalDomains": ["https://class2023.co.kr/"],
        "hospitalBlogBases": [
            "https://blog.naver.com/classlim2",
            "https://blog.naver.com/class231004",
        ],
    },
    "새로운병원": {
        "hospitalDomains": ["https://saerounhospital.com/"],
        "hospitalBlogBases": [
            "https://blog.naver.com/saerounhospital",
            "https://blog.naver.com/windyyarddd",
            "https://blog.naver.com/saerounhospital23",
        ],
        "hospitalCafeBases": [
            "https://cafe.naver.com/mktsesang",
            "https://cafe.naver.com/motiontree",
        ],
    },
    "대찬병원": {
        "hospitalDomains": ["https://www.daechanhospital.com/"],
        "hospitalBlogBases": [
            "https://blog.naver.com/dktk0525",
            "https://blog.naver.com/isongdos",
            "https://blog.naver.com/dchospital",
            "https://blog.naver.com/suca",
            "https://blog.naver.com/goals310",
        ],
        "hospitalCafeBases": [
            "https://cafe.naver.com/kyungmammo",
            "https://cafe.naver.com/tlgmdaka0",
        ],
    },
    "두발로병원": {
        "hospitalDomains": ["https://www.dubalo.com/"],
        "hospitalBlogBases": [
            "https://blog.naver.com/dubalo2021",
            "https://blog.naver.com/dubaloklee",
            "https://blog.naver.com/seoulfootandankle",
            "https://blog.naver.com/compare11529",
        ],
    },
    # 제로마취통증의학과: 공식 카페 없음 → 채점 시 카페 탭은 0점 고정(skipCafeScoring). 4월부터 별도 슬롯.
    "제로마취통증의학과": {
        "hospitalDomains": ["https://zeropainwonju.com/"],
        "hospitalBlogBases": ["https://blog.naver.com/zeropainwonju"],
        "skipCafeScoring": True,
    },
    "삼성본정형외과": {
        "hospitalDomains": ["https://www.samsungbon.com/"],
        "hospitalBlogBases": [
            "https://blog.naver.com/tight1530",
            "https://blog.naver.com/samsungbonhospital",
            "https://blog.naver.com/loveand0424",
        ],
        "hospitalCafeBases": [
            "https://cafe.naver.com/fo2b",
            "https://cafe.naver.com/bgga",
        ],
    },
    # 블로그목록/운영 표기에서 삼성본병원도 함께 쓰므로 동일 프로필로 취급
    "삼성본병원": {
        "hospitalDomains": ["https://www.samsungbon.com/"],
        "hospitalBlogBases": [
            "https://blog.naver.com/tight1530",
            "https://blog.naver.com/samsungbonhospital",
            "https://blog.naver.com/loveand0424",
        ],
        "hospitalCafeBases": [
            "https://cafe.naver.com/fo2b",
            "https://cafe.naver.com/bgga",
        ],
    },
    "제이엘정형외과": {
        "hospitalDomains": ["https://jlorthopedics.com/"],
        "hospitalBlogBases": [
            "https://blog.naver.com/muscle118072",
            "https://blog.naver.com/jlorthopedics",
        ],
        "hospitalCafeBases": [
            "https://cafe.naver.com/q5q5q5",
        ],
    },
    "SNU서울병원": {
        "hospitalDomains": ["https://snuseoul.kr/"],
        "hospitalBlogBases": [
            "https://blog.naver.com/snuseoulhospital",
            "https://blog.naver.com/dlwngks9808",
            "https://blog.naver.com/snuseoulfoot",
            "https://blog.naver.com/mission52814",
        ],
        "hospitalCafeBases": [
            "https://cafe.naver.com/q5q5q5",
            "https://cafe.naver.com/pajumom",
            "https://cafe.naver.com/usem",
        ],
    },
    "SNU건전비뇨의학과": {
        "hospitalDomains": ["https://www.seoulurology.com/"],
        "hospitalBlogBases": [
            "https://blog.naver.com/snugunjunpr",
            "https://blog.naver.com/secmet99",
        ],
    },
    "뉴민병원": {
        "hospitalDomains": ["https://www.newminhospital.com/"],
        "hospitalBlogBases": [
            "https://blog.naver.com/wealth_110412",
            "https://blog.naver.com/label_125212",
        ],
    },
    "더서울병원": {
        "hospitalDomains": ["https://www.theseoulhospital.com/"],
        "hospitalBlogBases": [
            "https://blog.naver.com/theseoulhospital",
            "https://blog.naver.com/brmhand",
            "https://blog.naver.com/veca2",
            "https://blog.naver.com/veca02",
        ],
        "hospitalCafeBases": [
            "https://cafe.naver.com/agameworld",
        ],
    },
    "방그레병원": {
        "hospitalDomains": ["https://bangre.co.kr/"],
        "hospitalBlogBases": [
            "https://blog.naver.com/kummd",
            "https://blog.naver.com/jdrcancer1109",
            "https://blog.naver.com/bangrenine",
        ],
    },
    "서울본정형외과": {
        "hospitalDomains": ["https://yonginseoulbone.com/"],
        "hospitalBlogBases": [
            "https://blog.naver.com/yonginseoulbone",
        ],
    },
    "서울센트럴병원": {
        "hospitalDomains": ["https://서울센트럴병원.com/"],
        "hospitalBlogBases": [
            "https://blog.naver.com/jhztuelsss18574",
            "https://blog.naver.com/seoulcentralhospital",
        ],
    },
    "서울현대정형외과": {
        "hospitalDomains": ["https://seoulhyundai.co.kr/"],
        "hospitalBlogBases": [
            "https://blog.naver.com/seoulhyundaihospital",
        ],
    },
    "서울현병원": {
        "hospitalDomains": ["https://seoulhyun.kr/"],
        "hospitalBlogBases": [
            "https://blog.naver.com/seoulhyunos",
            "https://blog.naver.com/sejongseoulhyun",
            "https://blog.naver.com/sejongoshyun",
            "https://blog.naver.com/zbxfzbmnr9091",
        ],
    },
    "신도림서울정형외과": {
        "hospitalDomains": ["https://www.sseoulos.com/"],
        "hospitalBlogBases": [
            "https://blog.naver.com/sseoulos",
        ],
    },
    "연세오케이병원": {
        "hospitalDomains": ["http://ysok.co.kr/orthopedics/"],
        "hospitalBlogBases": [
            "https://blog.naver.com/hanok0885",
            "https://blog.naver.com/owery1947",
            "https://blog.naver.com/ok0885ok",
        ],
    },
    "의정부항외과의원": {
        "hospitalDomains": ["http://hangsurgery.com/"],
        "hospitalBlogBases": [
            "https://blog.naver.com/ujbhang",
            "https://blog.naver.com/ujbkw",
        ],
    },
    "청주프라임병원": {
        "hospitalDomains": ["http://www.cjprime.com/"],
        "hospitalBlogBases": [
            "https://blog.naver.com/chungjuos",
            "https://blog.naver.com/cjprimestar14",
            "https://blog.naver.com/cjhealthsp",
        ],
    },
}


def available_hospitals_from_scoring() -> list[str]:
    """scoring-data.json 에 병합된 병원 + server 프로필이 있는 병원(채점 전 선택·업로드 가능)."""
    out: list[str] = ["포인트병원"]
    seen: set[str] = {"포인트병원"}
    if DATA_PATH.exists():
        try:
            data = json.loads(DATA_PATH.read_text(encoding="utf-8"))
            for m in data.get("months") or []:
                h = (m.get("hospitalName") or "").strip()
                if h and h not in seen:
                    seen.add(h)
                    out.append(h)
        except Exception:
            pass
    for h in HOSPITAL_PROFILE_OVERRIDES:
        if h and h not in seen:
            seen.add(h)
            out.append(h)
    for alias, canonical in HOSPITAL_CANONICAL.items():
        if canonical in seen and alias not in seen:
            seen.add(alias)
            out.append(alias)
    return out


# Live Server(5500)와 동일하게 index.html이 ./styles.css, ./app.js 를 루트에서 찾을 수 있게 함
app = Flask(__name__, static_folder=str(WEB_DIR), static_url_path="/web")
Compress(app)


@app.get("/ping")
def ping():
    return "", 204


@app.after_request
def add_cors_headers(resp):
    # 페이지를 다른 로컬 주소로 열어도 API 호출 가능하게 허용
    resp.headers["Access-Control-Allow-Origin"] = "*"
    resp.headers["Access-Control-Allow-Methods"] = "GET,POST,OPTIONS"
    resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
    # 스타일·스크립트·문서가 안 바뀌는 현상(브라우저 HTML 캐시 등) 줄이기
    if resp.mimetype in (
        "text/css",
        "text/javascript",
        "application/javascript",
        "text/html",
    ):
        resp.headers["Cache-Control"] = "no-store"
    return resp


def extract_keywords_from_text(text: str) -> list[str]:
    lines = [line.strip() for line in text.replace(",", "\n").splitlines()]
    seen = set()
    out = []
    for kw in lines:
        if not kw or kw in seen:
            continue
        seen.add(kw)
        out.append(kw)
    return out


def extract_keywords_from_xlsx(binary: bytes) -> list[str]:
    wb = load_workbook(io.BytesIO(binary), data_only=True)
    ws = wb.active
    keywords: list[str] = []
    seen = set()
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        # 첫 비어있지 않은 셀을 키워드로 간주
        first = next((c for c in row if c is not None and str(c).strip() != ""), None)
        if first is None:
            continue
        kw = str(first).strip()
        if kw.lower() in {"키워드", "keyword"}:
            continue
        if kw not in seen:
            seen.add(kw)
            keywords.append(kw)
    return keywords


def current_month_label() -> str:
    return f"{datetime.now().month}월"


def normalize_month_label(raw: str | None) -> str:
    s = (raw or "").strip()
    m = re.match(r"^(\d{1,2})월$", s)
    if not m:
        return current_month_label()
    n = int(m.group(1))
    if 1 <= n <= 12:
        return f"{n}월"
    return current_month_label()


def merge_keywords_keep_order(existing: list[str], incoming: list[str]) -> list[str]:
    seen = set()
    out = []
    for kw in (existing + incoming):
        k = (kw or "").strip()
        if not k or k in seen:
            continue
        seen.add(k)
        out.append(k)
    return out


def append_keywords_allow_duplicates(existing: list[str], incoming: list[str]) -> list[str]:
    """중복 허용: 기존 + 신규를 순서대로 연결(빈 문자열만 제거)."""
    out: list[str] = []
    for kw in existing:
        k = (kw or "").strip()
        if k:
            out.append(k)
    for kw in incoming:
        k = (kw or "").strip()
        if k:
            out.append(k)
    return out


def month_record_for_hospital(month_label: str, hospital_name: str) -> dict | None:
    """웹(app.js)의 monthRecordForHospitalMonth 와 동일한 우선순위.

    scoring-data.json 에 4월 등 동일 달에 hospitalName 없는 레거시(포인트)와
    hospitalName=\"포인트병원\" 인 항목이 같이 있으면, 예전 로직은 레거시를 먼저 골라
    키워드 수가 적은 쪽만 병합되어 추가가 반영되지 않는 문제가 있었다.
    """
    if not DATA_PATH.exists():
        return None
    try:
        root = json.loads(DATA_PATH.read_text(encoding="utf-8"))
        months = root.get("months") or []
        lab = (month_label or "").strip()
        raw_hn = (hospital_name or "").strip() or "포인트병원"

        if raw_hn == "포인트병원":
            for m in months:
                if (m.get("monthLabel") or "") != lab:
                    continue
                if str(m.get("hospitalName") or "").strip() == "포인트병원":
                    return m
            for m in months:
                if (m.get("monthLabel") or "") != lab:
                    continue
                raw = m.get("hospitalName")
                if raw is None or str(raw).strip() == "":
                    return m
            return None

        hn = canonical_hospital_name(hospital_name)
        for m in months:
            if (m.get("monthLabel") or "") != lab:
                continue
            if str(m.get("hospitalName") or "").strip() == hn:
                return m
    except Exception:
        return None
    return None


def load_keywords_for_month(month_label: str, hospital_name: str = "포인트병원") -> list[str]:
    # 1) scoring-data 에 이미 저장된 해당 월 키워드
    if DATA_PATH.exists():
        try:
            month = month_record_for_hospital(month_label, hospital_name)
            if month:
                sheets = month.get("sheets") or []
                if sheets:
                    kws: list[str] = []
                    for s in sheets:
                        for r in s.get("rows") or []:
                            kw = str((r[2] if len(r) > 2 else "") or "").strip()
                            if not kw:
                                continue
                            kws.append(kw)
                    if kws:
                        return kws
        except Exception:
            pass

    # 2) config 템플릿에 같은 월이 설정된 경우
    try:
        path = template_config_path(hospital_name)
        cfg = json.loads(path.read_text(encoding="utf-8"))
        if cfg.get("monthLabel") != month_label:
            return []
        if path.resolve() == CONFIG_PATH.resolve() and canonical_hospital_name(hospital_name) != "포인트병원":
            return []
        return [str(x).strip() for x in (cfg.get("keywords") or []) if str(x).strip()]
    except Exception:
        pass
    return []


def normalize_keyword_channel(raw: str | None) -> str:
    s = (raw or "all").strip().lower()
    return s if s in ("cafe", "blog", "all") else "all"


def normalize_keyword_scope(raw: str | None) -> str:
    s = (raw or "all").strip().lower()
    aliases = {
        "regional": "regional",
        "region": "regional",
        "지역": "regional",
        "national": "national",
        "전국": "national",
        "other": "other",
        "기타": "other",
        "etc": "other",
        "all": "all",
        "전체": "all",
    }
    return aliases.get(s, "all")


def sheet_keys_for_scope(scope: str) -> set[str]:
    s = normalize_keyword_scope(scope)
    all_keys = {k for k, _ in SHEET_ORDER}
    if s == "all":
        return all_keys
    if s == "regional":
        return {"regional-pc", "regional-mob"}
    if s == "national":
        return {"national-pc", "national-mob"}
    if s == "other":
        return {"other-pc", "other-mob"}
    return all_keys


def build_rows_by_sheet_key_for_month(
    month: dict[str, object] | None,
    merged: list[str],
    scopes: dict[str, str],
) -> tuple[dict[str, list[dict[str, str]]], dict[str, str]]:
    """키워드 범위(scope)에 따라 시트별 행 목록을 구성한다."""
    rows_by_sheet_key: dict[str, list[dict[str, str]]] = {}
    sheet_titles: dict[str, str] = {}
    old_by_key: dict[str, list[dict[str, str]]] = {}
    if month and isinstance(month.get("sheets"), list):
        for s in month["sheets"]:
            key = str(s.get("key") or "").strip()
            if not key:
                continue
            sheet_titles[key] = str(s.get("title") or key)
            pairs: list[dict[str, str]] = []
            for row in s.get("rows") or []:
                kw = str((row[2] if len(row) > 2 else "") or "").strip()
                if not kw:
                    continue
                reg = str((row[1] if len(row) > 1 else "") or "").strip()
                pairs.append({"region": reg, "keyword": kw})
            old_by_key[key] = pairs
    for key, default_title in SHEET_ORDER:
        if key not in sheet_titles:
            sheet_titles[key] = default_title
        old_pairs = old_by_key.get(key, [])
        reg_map: dict[str, str] = {}
        for p in old_pairs:
            kk = (p.get("keyword") or "").strip()
            if kk:
                reg_map[kk] = (p.get("region") or "").strip()
        pairs_out: list[dict[str, str]] = []
        for kw in merged:
            k = (kw or "").strip()
            if not k:
                continue
            sc = scopes.get(k) or "all"
            if key not in sheet_keys_for_scope(sc):
                continue
            pairs_out.append({"region": reg_map.get(k, ""), "keyword": k})
        rows_by_sheet_key[key] = pairs_out
    return rows_by_sheet_key, sheet_titles


def runtime_config_path_for_hospital(hospital_name: str) -> Path:
    digest = hashlib.sha1(canonical_hospital_name(hospital_name).encode("utf-8")).hexdigest()[:10]
    return ROOT / "config" / f"runtime_{digest}.json"


def update_keywords(
    keywords: list[str],
    month_label: str,
    hospital_name: str = "포인트병원",
    channel: str = "all",
    scope: str = "all",
) -> tuple[list[str], str, int]:
    hn = canonical_hospital_name(hospital_name)
    cfg = json.loads(template_config_path(hn).read_text(encoding="utf-8"))
    existing = load_keywords_for_month(month_label, hn)
    # 기존 목록에 행 단위 중복이 있어도 len(merged)-len(existing) 가 음수로 떨어지지 않게,
    # 이번 업로드에서 "처음 보는" 고유 키워드 수만 센다.
    existing_keys = {(kw or "").strip() for kw in existing if (kw or "").strip()}
    merged = merge_keywords_keep_order(existing, keywords)
    seen_incoming: set[str] = set()
    added_count = 0
    for kw in keywords:
        k = (kw or "").strip()
        if not k or k in seen_incoming:
            continue
        seen_incoming.add(k)
        if k not in existing_keys:
            added_count += 1
    cfg["monthLabel"] = month_label
    cfg["keywords"] = merged
    cfg["hospitalName"] = hn
    cfg["hospitalNames"] = [hn]
    for k, v in (HOSPITAL_PROFILE_OVERRIDES.get(hn) or {}).items():
        cfg[k] = v
    month = month_record_for_hospital(month_label, hn)
    ch = cfg.get("keywordChannels")
    if not isinstance(ch, dict):
        ch = {}
    if month and isinstance(month.get("keywordChannels"), dict):
        for k, v in month.get("keywordChannels").items():
            ks = str(k).strip()
            vs = normalize_keyword_channel(str(v))
            if ks:
                ch[ks] = vs
    norm = normalize_keyword_channel(channel)
    norm_scope = normalize_keyword_scope(scope)
    # 업로드 요청으로 들어온 키워드는 기존 동일월 데이터가 있어도 반드시 재채점 대상에 포함.
    force_rescore_keywords: list[str] = []
    seen_force: set[str] = set()
    for kw in keywords:
        k = (kw or "").strip()
        if k:
            ch[k] = norm
            if k not in seen_force:
                seen_force.add(k)
                force_rescore_keywords.append(k)
    cfg["keywordChannels"] = ch
    scopes = cfg.get("keywordScopes")
    if not isinstance(scopes, dict):
        scopes = {}
    if month and isinstance(month.get("keywordScopes"), dict):
        for k, v in month.get("keywordScopes").items():
            ks = str(k).strip()
            vs = normalize_keyword_scope(str(v))
            if ks:
                scopes[ks] = vs
    for kw in keywords:
        k = (kw or "").strip()
        if k:
            scopes[k] = norm_scope
    for kw in merged:
        k = (kw or "").strip()
        if k and k not in scopes:
            scopes[k] = "all"
    cfg["keywordScopes"] = scopes
    if month:
        rbk, st = build_rows_by_sheet_key_for_month(month, merged, scopes)
        cfg["rowsBySheetKey"] = rbk
        cfg["sheetTitles"] = st
    else:
        cfg.pop("rowsBySheetKey", None)
        cfg.pop("sheetTitles", None)
    cfg["forceRescoreKeywords"] = force_rescore_keywords
    out_path = runtime_config_path_for_hospital(hn)
    out_path.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
    return merged, out_path.name, added_count


def build_runtime_config_for_rerun(month_label: str, hospital_name: str) -> tuple[str | None, str]:
    """
    재채점 전용 런타임 config 생성.
    - scoring-data 의 선택 월/병원 데이터만 읽어 구성
    - 키워드 추가/삭제/변경은 하지 않음
    """
    hn = canonical_hospital_name(hospital_name)
    month = month_record_for_hospital(month_label, hn)
    if not month:
        return None, "선택한 월·병원 데이터가 없습니다."

    cfg = json.loads(template_config_path(hn).read_text(encoding="utf-8"))
    cfg["monthLabel"] = month_label
    cfg["hospitalName"] = hn
    cfg["hospitalNames"] = [hn]
    for k, v in (HOSPITAL_PROFILE_OVERRIDES.get(hn) or {}).items():
        cfg[k] = v

    flat_keywords: list[str] = []
    for s in month.get("sheets") or []:
        for row in s.get("rows") or []:
            kw = str((row[2] if len(row) > 2 else "") or "").strip()
            if kw:
                flat_keywords.append(kw)
    merged = merge_keywords_keep_order([], flat_keywords)
    if not merged:
        return None, "재채점할 키워드가 없습니다."

    month_kc = month.get("keywordChannels")
    if isinstance(month_kc, dict):
        cleaned: dict[str, str] = {}
        for k, v in month_kc.items():
            ks = str(k).strip()
            if not ks:
                continue
            cleaned[ks] = normalize_keyword_channel(str(v))
        cfg["keywordChannels"] = cleaned
    else:
        cfg["keywordChannels"] = {}
    scopes: dict[str, str] = {}
    month_ks = month.get("keywordScopes")
    if isinstance(month_ks, dict):
        for k, v in month_ks.items():
            ks = str(k).strip()
            if not ks:
                continue
            scopes[ks] = normalize_keyword_scope(str(v))
    for kw in merged:
        k = (kw or "").strip()
        if k and k not in scopes:
            scopes[k] = "all"
    cfg["keywordScopes"] = scopes

    rbk, st = build_rows_by_sheet_key_for_month(month, merged, scopes)
    cfg["keywords"] = merged
    cfg["rowsBySheetKey"] = rbk
    cfg["sheetTitles"] = st

    out_path = runtime_config_path_for_hospital(hn)
    out_path.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
    return out_path.name, ""


def delete_keyword(keyword: str) -> tuple[bool, int]:
    """모든 설정(config/runtime)에서 키워드 1개 삭제."""
    k = (keyword or "").strip()
    if not k:
        return False, 0
    changed = False
    remain_any = 0
    candidate_paths: list[Path] = [
        CONFIG_PATH,
        ZEROPAIN_CONFIG_PATH,
        SAMSUNGBON_CONFIG_PATH,
        JL_CONFIG_PATH,
        SNU_CONFIG_PATH,
    ]
    candidate_paths.extend(sorted((ROOT / "config").glob("runtime_*.json")))
    seen_paths: set[str] = set()
    for path in candidate_paths:
        key = str(path.resolve()) if path.exists() else str(path)
        if key in seen_paths:
            continue
        seen_paths.add(key)
        if not path.exists():
            continue
        try:
            cfg = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        current = [str(x).strip() for x in (cfg.get("keywords") or []) if str(x).strip()]
        if not current:
            continue
        next_keywords = [x for x in current if x != k]
        if len(next_keywords) != len(current):
            cfg["keywords"] = next_keywords
            ch = cfg.get("keywordChannels")
            if isinstance(ch, dict) and k in ch:
                ch.pop(k, None)
                cfg["keywordChannels"] = ch
            scopes = cfg.get("keywordScopes")
            if isinstance(scopes, dict) and k in scopes:
                scopes.pop(k, None)
                cfg["keywordScopes"] = scopes
            path.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
            changed = True
        remain_any = max(remain_any, len(next_keywords))
    return changed, remain_any


def delete_keyword_from_scoring_data(keyword: str) -> tuple[bool, int]:
    """scoring-data.json 모든 월/시트에서 키워드 행 제거."""
    k = (keyword or "").strip()
    if not k or not DATA_PATH.exists():
        return False, 0
    root = json.loads(DATA_PATH.read_text(encoding="utf-8"))
    changed = False
    removed = 0
    for month in root.get("months") or []:
        kc = month.get("keywordChannels")
        if isinstance(kc, dict) and k in kc:
            kc.pop(k, None)
            month["keywordChannels"] = kc
            changed = True
        ks = month.get("keywordScopes")
        if isinstance(ks, dict) and k in ks:
            ks.pop(k, None)
            month["keywordScopes"] = ks
            changed = True
        for sheet in month.get("sheets") or []:
            rows = sheet.get("rows") or []
            kept = []
            for row in rows:
                kw = str((row[2] if len(row) > 2 else "") or "").strip()
                if kw == k:
                    removed += 1
                    changed = True
                    continue
                kept.append(row)
            if len(kept) != len(rows):
                sheet["rows"] = kept
    if changed:
        DATA_PATH.write_text(json.dumps(root, ensure_ascii=False, indent=2), encoding="utf-8")
    return changed, removed


def delete_keyword_from_evidence(keyword: str) -> bool:
    """last-run-evidence.json 에서 키워드 근거 제거."""
    k = (keyword or "").strip()
    ev_path = ROOT / "data" / "last-run-evidence.json"
    if not k or not ev_path.exists():
        return False
    root = json.loads(ev_path.read_text(encoding="utf-8"))
    changed = False
    ev = root.get("evidence")
    if isinstance(ev, dict) and k in ev:
        ev.pop(k, None)
        root["evidence"] = ev
        changed = True
    by_h = root.get("byHospital")
    if isinstance(by_h, dict):
        for name, obj in by_h.items():
            if isinstance(obj, dict) and k in obj:
                obj.pop(k, None)
                by_h[name] = obj
                changed = True
        root["byHospital"] = by_h
    if changed:
        ev_path.write_text(json.dumps(root, ensure_ascii=False, indent=2), encoding="utf-8")
    return changed


def _github_push_scoring_files() -> None:
    """채점 완료 후 scoring-data.json + last-run-evidence.json 을 GitHub에 자동 커밋.
    Git Data API 사용 → 대용량 파일(1MB+) 처리 가능."""
    token = os.getenv("GITHUB_TOKEN", "").strip()
    if not token:
        return
    owner = os.getenv("GITHUB_REPO_OWNER", "chadesign0").strip()
    repo  = os.getenv("GITHUB_REPO_NAME",  "vibe-coding_-study2").strip()

    files: list[Path] = []
    if DATA_PATH.exists():
        files.append(DATA_PATH)
    if not files:
        return

    headers = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github+json",
        "Content-Type": "application/json",
    }
    base = f"https://api.github.com/repos/{owner}/{repo}"

    def gh(url: str, data: dict | None = None, method: str | None = None) -> dict:
        body = json.dumps(data).encode() if data is not None else None
        req = urllib.request.Request(
            url, data=body, headers=headers,
            method=method or ("POST" if body else "GET"),
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            return json.loads(resp.read())

    try:
        # 1. 각 파일 blob 생성
        tree_entries = []
        for fp in files:
            blob = gh(f"{base}/git/blobs", {
                "content": base64.b64encode(fp.read_bytes()).decode(),
                "encoding": "base64",
            })
            tree_entries.append({
                "path": fp.relative_to(ROOT).as_posix(),
                "mode": "100644", "type": "blob", "sha": blob["sha"],
            })

        # 2. 현재 main 커밋 SHA
        ref = gh(f"{base}/git/refs/heads/main")
        latest_sha = ref["object"]["sha"]

        # 3. 현재 tree SHA
        tree_sha = gh(f"{base}/git/commits/{latest_sha}")["tree"]["sha"]

        # 4. 새 tree 생성
        new_tree_sha = gh(f"{base}/git/trees", {
            "base_tree": tree_sha, "tree": tree_entries,
        })["sha"]

        # 5. 커밋 생성
        from datetime import datetime as _dt
        now = _dt.now().strftime("%Y-%m-%d %H:%M")
        new_sha = gh(f"{base}/git/commits", {
            "message": f"auto: 채점 데이터 업데이트 ({now})",
            "tree": new_tree_sha,
            "parents": [latest_sha],
        })["sha"]

        # 6. main 브랜치 업데이트
        gh(f"{base}/git/refs/heads/main", {"sha": new_sha}, method="PATCH")
        print(f"[github] 자동 커밋 성공: {new_sha[:7]}")

    except Exception as e:
        print(f"[github] 커밋 실패: {e}")


def _kakao_notify(text: str) -> None:
    """카카오톡 나에게 보내기. KAKAO_ACCESS_TOKEN 환경변수 미설정 시 무시."""
    access_token = os.getenv("KAKAO_ACCESS_TOKEN", "").strip()
    if not access_token:
        return
    template = json.dumps(
        {"object_type": "text", "text": text, "link": {"web_url": "https://baejeompyojadonghwa.onrender.com/"}},
        ensure_ascii=False,
    )
    data = urllib.parse.urlencode({"template_object": template}).encode()
    req = urllib.request.Request(
        "https://kapi.kakao.com/v2/api/talk/memo/default/send",
        data=data,
        headers={"Authorization": f"Bearer {access_token}"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            result = json.loads(resp.read())
            if result.get("result_code") != 0:
                print(f"[kakao] 전송 실패: {result}")
    except Exception as e:
        print(f"[kakao] 오류: {e}")


def _check_data_size_and_notify() -> None:
    """scoring-data.json 이 500KB 초과 시 카카오톡 알림."""
    if not DATA_PATH.exists():
        return
    size_kb = DATA_PATH.stat().st_size / 1024
    if size_kb > 500:
        msg = (
            f"[배점표 자동화 알림]\n\n"
            f"⚠️ 데이터 용량 경고\n\n"
            f"scoring-data.json 이 500KB를 초과했습니다.\n"
            f"현재 크기: {size_kb:.0f}KB\n\n"
            f"MongoDB 전환을 검토할 시점입니다."
        )
        _kakao_notify(msg)


def _merge_scoring_temp(temp_path: Path) -> None:
    """임시 채점 결과를 scoring-data.json 에 안전하게 병합 (MERGE_LOCK 안에서 호출)."""
    month = json.loads(temp_path.read_text(encoding="utf-8"))
    root = json.loads(DATA_PATH.read_text(encoding="utf-8")) if DATA_PATH.exists() else {"months": []}
    new_lab = str(month.get("monthLabel") or "").strip()
    new_hn = str(month.get("hospitalName") or "").strip()

    def _identity(m: dict) -> tuple:
        return (str(m.get("monthLabel") or "").strip(), str(m.get("hospitalName") or "").strip())

    new_id = (new_lab, new_hn)

    def _keep(m: dict) -> bool:
        if _identity(m) == new_id:
            return False
        if new_hn == "포인트병원" and new_lab:
            ml = str(m.get("monthLabel") or "").strip()
            raw = m.get("hospitalName")
            if ml == new_lab and (raw is None or str(raw).strip() == ""):
                return False
        return True

    def _month_order(m: dict) -> int:
        import re as _re
        match = _re.match(r"^(\d{1,2})월\s*$", str(m.get("monthLabel") or "").strip())
        return int(match.group(1)) if match else 99

    months = [m for m in (root.get("months") or []) if _keep(m)]
    months.append(month)
    months.sort(key=_month_order)
    root["months"] = months
    root["generatedBy"] = "build_april_month.py(api+web-evidence)"
    DATA_PATH.write_text(json.dumps(root, ensure_ascii=False, indent=2), encoding="utf-8")


def _merge_evidence_temp(temp_path: Path) -> None:
    """임시 근거 결과를 last-run-evidence.json 에 안전하게 병합 (MERGE_LOCK 안에서 호출)."""
    import time as _time
    payload = json.loads(temp_path.read_text(encoding="utf-8"))
    evidence = payload.get("evidence") or {}
    hospital_name = payload.get("hospitalName") or None
    ev_path = ROOT / "data" / "last-run-evidence.json"
    flat: dict = {}
    by_h: dict = {}
    if ev_path.exists():
        try:
            old = json.loads(ev_path.read_text(encoding="utf-8"))
            flat = dict(old.get("evidence") or {})
            by_h = dict(old.get("byHospital") or {})
        except Exception:
            pass
    if hospital_name:
        key = str(hospital_name).strip()
        merged_h = dict(by_h.get(key) or {})
        merged_h.update(evidence or {})
        by_h[key] = merged_h
    else:
        flat = evidence
    result = {"generatedAt": _time.strftime("%Y-%m-%d %H:%M:%S"), "evidence": flat, "byHospital": by_h}
    ev_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")


def run_scoring(config_name: str | None = None, *, full_rescore: bool = False) -> tuple[bool, str]:
    task_uid = uuid.uuid4().hex
    temp_scoring = ROOT / "data" / f"_tmp_scoring_{task_uid}.json"
    temp_evidence = ROOT / "data" / f"_tmp_evidence_{task_uid}.json"

    env = os.environ.copy()
    if config_name:
        env["SCORING_CONFIG"] = config_name
    if full_rescore:
        env["SCORING_FULL_RESCORE"] = "1"
    else:
        env.pop("SCORING_FULL_RESCORE", None)
    env["SCORING_TEMP_OUTPUT"] = str(temp_scoring)
    env["EVIDENCE_TEMP_OUTPUT"] = str(temp_evidence)

    proc = subprocess.run(
        ["python", str(SCRIPT_PATH)],
        cwd=str(ROOT),
        env=env,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    out = (proc.stdout or "") + (proc.stderr or "")

    if proc.returncode == 0:
        with MERGE_LOCK:
            try:
                if temp_scoring.exists():
                    _merge_scoring_temp(temp_scoring)
                if temp_evidence.exists():
                    _merge_evidence_temp(temp_evidence)
            finally:
                temp_scoring.unlink(missing_ok=True)
                temp_evidence.unlink(missing_ok=True)
        threading.Thread(target=_github_push_scoring_files, daemon=True).start()
        threading.Thread(target=_check_data_size_and_notify, daemon=True).start()
    else:
        temp_scoring.unlink(missing_ok=True)
        temp_evidence.unlink(missing_ok=True)

    return proc.returncode == 0, out


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _task_key(kind: str, hospital_name: str, month_label: str) -> str:
    return f"{kind}:{hospital_name}:{month_label}"


def get_score_task(task_id: str) -> dict[str, object] | None:
    with SCORE_TASKS_LOCK:
        raw = SCORE_TASKS.get(task_id)
        if not isinstance(raw, dict):
            return None
        out = dict(raw)
        expose_log = (os.getenv("SCORING_EXPOSE_LOG") or "").strip().lower() in {"1", "true", "yes"}
        if not expose_log:
            out.pop("log", None)
        return out


def _set_task_status(task_id: str, **patch: object) -> None:
    with SCORE_TASKS_LOCK:
        cur = SCORE_TASKS.get(task_id)
        if not isinstance(cur, dict):
            return
        cur.update(patch)
        SCORE_TASKS[task_id] = cur


def enqueue_score_task(
    *,
    kind: str,
    hospital_name: str,
    month_label: str,
    config_name: str,
    message: str,
    meta: dict[str, object] | None = None,
    full_rescore: bool = False,
) -> tuple[str, bool]:
    """
    비동기 채점 작업 등록/실행.
    반환: (task_id, created_new)
    - 같은 병원/월/작업종류에서 기존 queued/running 이 있으면 그 task_id 재사용.
    """
    key = _task_key(kind, hospital_name, month_label)
    with SCORE_TASKS_LOCK:
        active_id = ACTIVE_SCORE_TASK_BY_KEY.get(key)
        if active_id:
            existing = SCORE_TASKS.get(active_id) or {}
            if existing.get("status") in {"queued", "running"}:
                return active_id, False

        task_id = uuid.uuid4().hex
        task: dict[str, object] = {
            "taskId": task_id,
            "kind": kind,
            "status": "queued",
            "hospitalName": hospital_name,
            "monthLabel": month_label,
            "message": message,
            "configName": config_name,
            "fullRescore": full_rescore,
            "createdAt": now_iso(),
            "startedAt": None,
            "endedAt": None,
            "log": "",
            "error": "",
        }
        if isinstance(meta, dict):
            task["meta"] = meta
        SCORE_TASKS[task_id] = task
        ACTIVE_SCORE_TASK_BY_KEY[key] = task_id

    def _worker() -> None:
        _set_task_status(task_id, status="running", startedAt=now_iso())
        ok, log = run_scoring(config_name, full_rescore=full_rescore)
        trimmed = (log or "")[-120000:]
        q_level = None
        q_acc = None
        m = re.search(r"QUALITY_GATE:(OK|WARN|LOW):([0-9]+(?:\.[0-9]+)?)", trimmed)
        if m:
            q_level = m.group(1).lower()
            try:
                q_acc = float(m.group(2))
            except Exception:
                q_acc = None
        quality_payload = {"level": q_level, "accuracyPct": q_acc} if q_level else None
        if ok:
            _set_task_status(
                task_id,
                status="succeeded",
                endedAt=now_iso(),
                log=trimmed,
                message=(
                    f"{hospital_name} · {month_label} 채점 완료"
                    if q_level not in {"warn", "low"}
                    else f"{hospital_name} · {month_label} 채점 완료 (정확도 경고)"
                ),
                quality=quality_payload,
            )
        else:
            gate_msg = "채점 실행 실패"
            _set_task_status(
                task_id,
                status="failed",
                endedAt=now_iso(),
                log=trimmed,
                error=gate_msg,
                message=f"{hospital_name} · {month_label} {gate_msg}",
                quality=quality_payload,
            )
        with SCORE_TASKS_LOCK:
            if ACTIVE_SCORE_TASK_BY_KEY.get(key) == task_id:
                ACTIVE_SCORE_TASK_BY_KEY.pop(key, None)

    threading.Thread(target=_worker, daemon=True, name=f"score-task-{task_id[:8]}").start()
    return task_id, True


def load_hospital_list() -> list[str]:
    if not HOSPITAL_LIST_PATH.exists():
        return ["포인트병원"]
    seen = set()
    out: list[str] = []
    for raw in HOSPITAL_LIST_PATH.read_text(encoding="utf-8").splitlines():
        t = raw.strip()
        if not t or t.startswith("#"):
            continue
        if t in seen:
            continue
        seen.add(t)
        out.append(t)
    return out or ["포인트병원"]


@app.get("/")
def root():
    return send_from_directory(str(WEB_DIR), "index.html")


@app.get("/styles.css")
def serve_styles():
    return send_file(WEB_DIR / "styles.css", mimetype="text/css", max_age=0)


@app.get("/app.js")
def serve_app_js():
    return send_file(WEB_DIR / "app.js", mimetype="text/javascript", max_age=0)



KNOWN_FONT_FILES = (
    "MagdaCleanMono-Regular.otf",
    "YangJin.otf",
    "SCDream3.otf",
)


def _font_search_bases() -> list[Path]:
    """프로젝트·web 아래 흔한 폴더명과 web 루트(플랫)까지 순서대로 탐색."""
    return [
        WEB_DIR / "@font",
        WEB_DIR / "font",
        WEB_DIR / "fonts",
        ROOT / "@font",
        ROOT / "font",
        ROOT / "fonts",
        WEB_DIR,
    ]


def _find_font_on_disk(filename: str) -> Path | None:
    """이름만 일치하는 폰트 파일 경로 (프로젝트 루트 밖으로는 나가지 않음)."""
    safe = filename.replace("\\", "/").split("/")[-1]
    if not safe or ".." in safe or "/" in filename.replace("\\", "/"):
        return None
    ext = Path(safe).suffix.lower()
    if ext not in (".otf", ".ttf", ".woff", ".woff2"):
        return None
    try:
        root_bound = ROOT.resolve()
    except OSError:
        return None
    for base in _font_search_bases():
        try:
            if not base.is_dir():
                continue
        except OSError:
            continue
        candidate = (base / safe).resolve()
        if not candidate.is_file():
            continue
        try:
            candidate.relative_to(root_bound)
        except ValueError:
            continue
        return candidate
    return None


def _send_font_candidate(candidate: Path):
    ext = candidate.suffix.lower()
    mimetype = {
        ".otf": "font/otf",
        ".ttf": "font/ttf",
        ".woff2": "font/woff2",
        ".woff": "font/woff",
    }.get(ext, "application/octet-stream")
    resp = send_file(candidate, mimetype=mimetype, max_age=0)
    resp.headers["Cache-Control"] = "no-store"
    return resp


def _serve_font_by_name(filename: str):
    """styles.css와 같은 루트 기준 상대 경로(url('파일.otf')) + /font/파일명 모두 지원."""
    candidate = _find_font_on_disk(filename)
    if candidate is None:
        abort(404)
    return _send_font_candidate(candidate)


@app.get("/font/<filename>")
def serve_font_file(filename: str):
    """예: /font/MagdaCleanMono-Regular.otf"""
    return _serve_font_by_name(filename)


def _register_known_fonts_at_root():
    """CSS 상대 url('…otf') → 브라우저가 /파일명.otf 로 요청할 때 매칭."""

    def _make_handler(name: str):
        def _view():
            return _serve_font_by_name(name)

        _view.__name__ = f"serve_font_root_{name.replace('.', '_').replace('-', '_')}"
        return _view

    for i, fname in enumerate(KNOWN_FONT_FILES):
        app.add_url_rule(
            f"/{fname}",
            view_func=_make_handler(fname),
            endpoint=f"font_at_root_{i}",
        )


_register_known_fonts_at_root()


@app.post("/api/upload-keywords")
def upload_keywords():
    text_input = (request.form.get("keywords_text") or "").strip()
    file = request.files.get("keywords_file")
    keywords: list[str] = []

    if text_input:
        keywords = extract_keywords_from_text(text_input)
    elif file:
        name = (file.filename or "").lower()
        payload = file.read()
        if name.endswith(".xlsx"):
            keywords = extract_keywords_from_xlsx(payload)
        else:
            keywords = extract_keywords_from_text(payload.decode("utf-8", errors="ignore"))
    else:
        return jsonify({"ok": False, "message": "텍스트 또는 파일을 입력해주세요."}), 400

    if not keywords:
        return jsonify({"ok": False, "message": "유효한 키워드를 찾지 못했습니다."}), 400

    month_label = normalize_month_label(request.form.get("month_label"))
    hospital_name = canonical_hospital_name(
        (request.form.get("hospital_name") or "포인트병원").strip() or "포인트병원"
    )
    channel = normalize_keyword_channel(request.form.get("keyword_channel"))
    scope = normalize_keyword_scope(request.form.get("keyword_scope"))
    merged, config_name, added_count = update_keywords(keywords, month_label, hospital_name, channel, scope)
    force_sync = (request.form.get("sync") or "").strip().lower() in {"1", "true", "yes"}
    if force_sync:
        ok, log = run_scoring(config_name, full_rescore=False)
        if not ok:
            return jsonify({"ok": False, "message": "채점 실행 실패"}), 500
        return jsonify(
            {
                "ok": True,
                "accepted": False,
                "message": f"{hospital_name} · {month_label} 키워드 업로드 및 자동 채점 완료",
                "count": len(merged),
                "addedCount": added_count,
                "monthLabel": month_label,
                "hospitalName": hospital_name,
            }
        )
    task_id, created_new = enqueue_score_task(
        kind="upload_keywords",
        hospital_name=hospital_name,
        month_label=month_label,
        config_name=config_name,
        message=f"{hospital_name} · {month_label} 키워드 채점 대기중",
        meta={"count": len(merged)},
        full_rescore=False,
    )
    return jsonify(
        {
            "ok": True,
            "accepted": True,
            "taskId": task_id,
            "createdNewTask": created_new,
            "message": f"{hospital_name} · {month_label} 키워드 채점 작업을 시작했습니다.",
            "count": len(merged),
            "addedCount": added_count,
            "monthLabel": month_label,
            "hospitalName": hospital_name,
        }
    ), 202


@app.post("/api/delete-keyword")
def delete_keyword_api():
    keyword = (request.form.get("keyword") or "").strip()
    if not keyword:
        return jsonify({"ok": False, "message": "삭제할 키워드를 입력해주세요."}), 400
    cfg_deleted, remain = delete_keyword(keyword)
    data_deleted, removed_rows = delete_keyword_from_scoring_data(keyword)
    ev_deleted = delete_keyword_from_evidence(keyword)
    if not (cfg_deleted or data_deleted or ev_deleted):
        return jsonify({"ok": False, "message": "해당 키워드를 찾지 못했습니다."}), 404
    return jsonify(
        {
            "ok": True,
            "message": f"키워드 삭제 완료: {keyword}",
            "remaining": remain,
            "removedRows": removed_rows,
        }
    )


@app.post("/api/run-score")
def run_score():
    hospital_name = canonical_hospital_name(
        (request.form.get("hospital_name") or "포인트병원").strip() or "포인트병원"
    )
    month_label = normalize_month_label(request.form.get("month_label"))
    config_name, err = build_runtime_config_for_rerun(month_label, hospital_name)
    if not config_name:
        return jsonify({"ok": False, "message": err or "재채점 준비 실패"}), 400
    force_sync = (request.form.get("sync") or "").strip().lower() in {"1", "true", "yes"}
    if force_sync:
        ok, log = run_scoring(config_name, full_rescore=True)
        return jsonify({"ok": ok, "accepted": False}), (200 if ok else 500)
    task_id, created_new = enqueue_score_task(
        kind="rerun_score",
        hospital_name=hospital_name,
        month_label=month_label,
        config_name=config_name,
        message=f"{hospital_name} · {month_label} 재채점 대기중",
        meta={},
        full_rescore=True,
    )
    return jsonify({"ok": True, "accepted": True, "taskId": task_id, "createdNewTask": created_new}), 202


@app.get("/api/score-task/<task_id>")
def score_task_status(task_id: str):
    task = get_score_task(task_id)
    if not task:
        return jsonify({"ok": False, "message": "작업을 찾을 수 없습니다."}), 404
    return jsonify({"ok": True, "task": task})


@app.get("/api/hospitals")
def hospitals():
    items = load_hospital_list()
    available = available_hospitals_from_scoring()
    return jsonify({"ok": True, "hospitals": items, "availableHospitals": available})


@app.post("/api/export-xlsx")
def export_xlsx():
    payload = request.get_json(silent=True) or {}
    headers = payload.get("headers") or []
    rows = payload.get("rows") or []
    file_name_raw = str(payload.get("fileName") or "").strip() or "score-export"
    safe_name = re.sub(r"[^\w가-힣\-\.\(\) ]+", "_", file_name_raw).strip() or "score-export"
    if not safe_name.lower().endswith(".xlsx"):
        safe_name += ".xlsx"
    if not isinstance(headers, list) or not headers:
        return jsonify({"ok": False, "message": "내보낼 헤더가 없습니다."}), 400
    if not isinstance(rows, list):
        return jsonify({"ok": False, "message": "내보낼 행 형식이 올바르지 않습니다."}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "배점표"
    ws.append([str(x) for x in headers])
    for row in rows:
        if not isinstance(row, list):
            continue
        ws.append([("" if x is None else str(x)) for x in row[: len(headers)]])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=safe_name,
        max_age=0,
    )


@app.get("/data/<path:filename>")
def data_files(filename: str):
    target = ROOT / "data" / filename
    if not target.exists() and filename == "last-run-evidence.json":
        resp = jsonify({"evidence": {}})
        resp.headers["Cache-Control"] = "no-store"
        return resp, 200
    resp = send_from_directory(str(ROOT / "data"), filename)
    resp.headers["Cache-Control"] = "no-store"
    return resp


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port, debug=False)

