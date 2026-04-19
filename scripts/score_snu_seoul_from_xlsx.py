# -*- coding: utf-8 -*-
"""
SNU서울병원 키워드 엑셀 → config/snu_march_keywords.json 저장 후 네이버 API 채점·scoring-data.json 병합.

- SNU_KEYWORD_XLSX: 엑셀 경로 (기본: 월간보고서 202603 파일)
- SNU_SKIP_SCORING=1 이면 JSON만 저장하고 채점은 건너뜀
"""
from __future__ import annotations

import importlib.util
import json
import os
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
CONFIG_OUT = ROOT / "config" / "snu_march_keywords.json"

DEFAULT_XLSX = (
    r"y:\@컨텐츠팀\S_SNU서울병원\01. 월간보고서\2026년\202603\SNU서울병원_배점표_26년 03월.xlsx"
)
SHEET_KEYS = ("regional-pc", "regional-mob", "national-pc", "national-mob")


def _load_build_module():
    path = ROOT / "scripts" / "build_april_month.py"
    spec = importlib.util.spec_from_file_location("build_april_month", path)
    if spec is None or spec.loader is None:
        raise RuntimeError("build_april_month 로드 실패")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def resolve_month_label(xlsx: Path) -> str:
    env = (os.getenv("SNU_MONTH_LABEL") or "").strip()
    if env:
        return env
    m = re.search(r"(\d{1,2})월", xlsx.name)
    if m:
        return f"{int(m.group(1))}월"
    return "3월"


def sheet_key_from_excel_name(name: str) -> str | None:
    n = (name or "").replace(" ", "").strip()
    if not n:
        return None
    is_mob = ("모바일" in n) or ("MOB" in n.upper())
    scope = "national" if "전국" in n else "regional"
    return f"{scope}-{'mob' if is_mob else 'pc'}"


def find_header_and_cols(ws) -> tuple[int, int, int] | None:
    max_r = min(ws.max_row or 0, 40)
    for r in range(1, max_r + 1):
        reg_c = kw_c = None
        for c in range(1, min(ws.max_column or 0, 40) + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).replace(" ", "").strip()
            if s == "지역":
                reg_c = c
            if "키워드" in s and "연관" not in s and "검색어" not in s and "합계" not in s:
                kw_c = c
        if reg_c and kw_c:
            return r, reg_c, kw_c
    return None


def read_pairs_keyword_only(ws) -> list[dict[str, str]]:
    found = find_header_and_cols(ws)
    if not found:
        return []
    header_r, reg_c, kw_c = found
    out: list[dict[str, str]] = []
    for r in range(header_r + 1, (ws.max_row or 0) + 1):
        raw_kw = ws.cell(r, kw_c).value
        if raw_kw is None or str(raw_kw).strip() == "":
            continue
        kw = str(raw_kw).strip()
        raw_reg = ws.cell(r, reg_c).value
        reg = str(raw_reg).strip() if raw_reg is not None else ""
        out.append({"region": reg, "keyword": kw})
    return out


def build_config_from_xlsx(xlsx: Path) -> dict[str, Any]:
    wb = load_workbook(xlsx, read_only=False, data_only=True)
    rows_by: dict[str, list[dict[str, str]]] = {k: [] for k in SHEET_KEYS}
    for raw_name in wb.sheetnames:
        key = sheet_key_from_excel_name(raw_name)
        if key not in SHEET_KEYS:
            continue
        ws = wb[raw_name]
        pairs = read_pairs_keyword_only(ws)
        if not pairs:
            continue
        rows_by[key].extend(pairs)
    wb.close()

    for key in SHEET_KEYS:
        dedup: list[dict[str, str]] = []
        seen: set[str] = set()
        for p in rows_by.get(key) or []:
            kw = (p.get("keyword") or "").strip()
            if not kw or kw in seen:
                continue
            seen.add(kw)
            dedup.append({"region": (p.get("region") or "").strip(), "keyword": kw})
        rows_by[key] = dedup

    missing = [k for k in SHEET_KEYS if not rows_by.get(k)]
    if missing:
        raise SystemExit(f"필수 시트가 없습니다: {missing}. 현재: {list(rows_by.keys())}")

    all_kw: list[str] = []
    seen: set[str] = set()
    for key in SHEET_KEYS:
        for p in rows_by[key]:
            k = p["keyword"]
            if k not in seen:
                seen.add(k)
                all_kw.append(k)

    month_label = resolve_month_label(xlsx)
    return {
        "monthLabel": month_label,
        "scoringYear": 2026,
        "blogRequirePostMonth": True,
        "hospitalName": "SNU서울병원",
        "hospitalNames": ["SNU서울병원"],
        "hospitalDomains": ["https://snuseoul.kr/"],
        "hospitalBlogBases": [
            "https://blog.naver.com/snuseoulhospital",
            "https://blog.naver.com/dlwngks9808",
            "https://blog.naver.com/snuseoulfoot",
            "https://blog.naver.com/mission52814",
        ],
        "hospitalCafeBases": [
            "https://cafe.naver.com/q5q5q5",
            "https://cafe.naver.com/pajumom",
            "https://cafe.naver.com/usem",
        ],
        "sourceFileNote": str(xlsx.name),
        "manualRanks": {},
        "manualRanksByTab": {},
        "regionDefault": "",
        "keywords": all_kw,
        "rowsBySheetKey": rows_by,
        "sheetTitles": {
            "regional-pc": "2026 지역 PC",
            "regional-mob": "2026 지역 모바일",
            "national-pc": "2026 전국 PC",
            "national-mob": "2026 전국 모바일",
        },
    }


def main() -> None:
    xlsx = Path((os.getenv("SNU_KEYWORD_XLSX") or DEFAULT_XLSX).strip())
    if not xlsx.exists():
        raise SystemExit(f"엑셀을 찾을 수 없습니다: {xlsx}")
    ml = resolve_month_label(xlsx)
    print("병합 슬롯: monthLabel =", ml, "| hospitalName = SNU서울병원")
    cfg = build_config_from_xlsx(xlsx)
    print("키워드 수(중복 제거):", len(cfg["keywords"]))
    print("시트별 행 수:", {k: len(v) for k, v in cfg["rowsBySheetKey"].items()})

    CONFIG_OUT.parent.mkdir(parents=True, exist_ok=True)
    CONFIG_OUT.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
    print("저장:", CONFIG_OUT)

    if (os.getenv("SNU_SKIP_SCORING") or "").strip() in ("1", "true", "yes"):
        print("SNU_SKIP_SCORING: skip API scoring")
        return

    bm = _load_build_module()
    bm.run_scoring_pipeline(cfg)
    print("완료: scoring-data.json 병합, last-run-evidence.json 의 byHospital['SNU서울병원'] 갱신")


if __name__ == "__main__":
    main()
