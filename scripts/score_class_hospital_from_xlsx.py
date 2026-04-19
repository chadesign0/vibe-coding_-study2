# -*- coding: utf-8 -*-
"""
클래스병원 키워드 엑셀: 지역·키워드 열만 읽고 네이버 API로 재채점 후 scoring-data.json 에 병합.

- 포인트병원 월 데이터와는 (monthLabel, hospitalName) 슬롯이 달라 서로 덮어쓰지 않음.
- 엑셀 경로: CLASS_KEYWORD_XLSX (기본값은 3월 예시 경로)
- 월 라벨: CLASS_MONTH_LABEL (예: 4월) 또는 파일명에 「N월」이 있으면 자동 추론

실행(프로젝트 루트, NAVER_CLIENT_ID / NAVER_CLIENT_SECRET 필요):
  python scripts/score_class_hospital_from_xlsx.py
  set CLASS_MONTH_LABEL=4월 && python scripts/score_class_hospital_from_xlsx.py
"""
from __future__ import annotations

import importlib.util
import os
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]


def _load_build_module():
    path = ROOT / "scripts" / "build_april_month.py"
    spec = importlib.util.spec_from_file_location("build_april_month", path)
    if spec is None or spec.loader is None:
        raise RuntimeError("build_april_month 로드 실패")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod

DEFAULT_XLSX = r"y:\@컨텐츠팀\C_클래스병원\02.월간보고서\3월\클래스병원키워드_키워드배점표_3월.xlsx"


def resolve_month_label(xlsx: Path) -> str:
    """CLASS_MONTH_LABEL 우선, 없으면 파일명의 N월, 기본 3월."""
    env = (os.getenv("CLASS_MONTH_LABEL") or "").strip()
    if env:
        return env
    m = re.search(r"(\d{1,2})월", xlsx.name)
    if m:
        return f"{int(m.group(1))}월"
    return "3월"

SHEET_KEYS = ("regional-pc", "regional-mob", "national-pc", "national-mob")


def sheet_key_from_excel_name(name: str) -> str | None:
    n = (name or "").strip()
    if "지역" in n and ("모바일" in n or "MOB" in n.upper()):
        return "regional-mob"
    if "지역" in n and "PC" in n.upper():
        return "regional-pc"
    if "전국" in n and ("모바일" in n or "MOB" in n.upper()):
        return "national-mob"
    if "전국" in n and "PC" in n.upper():
        return "national-pc"
    return None


def find_header_and_cols(ws) -> tuple[int, int, int] | None:
    """(header_row, region_col, keyword_col) 1-based. 실패 시 None."""
    max_r = min(ws.max_row or 0, 40)
    for r in range(1, max_r + 1):
        reg_c = kw_c = None
        for c in range(1, min(ws.max_column or 0, 40) + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).replace(" ", "").strip()
            if s == "지역":
                reg_c = c
            # '키워드별 합계' 등은 제외(마지막 열이 잡히면 번호·점수만 읽는 오류 발생)
            if "키워드" in s and "연관" not in s and "검색어" not in s and "합계" not in s:
                kw_c = c
        if reg_c and kw_c:
            return r, reg_c, kw_c
    return None


def read_pairs_keyword_only(ws) -> list[dict[str, str]]:
    """헤더 아래부터 지역·키워드 열만 사용. 나머지 열은 읽지 않음."""
    found = find_header_and_cols(ws)
    if not found:
        return []
    header_r, reg_c, kw_c = found
    out: list[dict[str, str]] = []
    for r in range(header_r + 1, (ws.max_row or 0) + 1):
        raw_kw = ws.cell(r, kw_c).value
        if raw_kw is None or str(raw_kw).strip() == "":
            continue
        kw = str(raw_kw).strip()
        raw_reg = ws.cell(r, reg_c).value
        reg = str(raw_reg).strip() if raw_reg is not None else ""
        out.append({"region": reg, "keyword": kw})
    return out


def build_config_from_xlsx(xlsx: Path) -> dict[str, Any]:
    wb = load_workbook(xlsx, read_only=False, data_only=True)
    rows_by: dict[str, list[dict[str, str]]] = {}
    for raw_name in wb.sheetnames:
        key = sheet_key_from_excel_name(raw_name)
        if key not in SHEET_KEYS:
            continue
        ws = wb[raw_name]
        pairs = read_pairs_keyword_only(ws)
        if not pairs:
            raise SystemExit(f"시트 '{raw_name}' 에서 지역/키워드 헤더를 찾지 못했습니다.")
        rows_by[key] = pairs
    wb.close()

    missing = [k for k in SHEET_KEYS if k not in rows_by]
    if missing:
        raise SystemExit(f"필수 시트가 없습니다: {missing}. 현재: {list(rows_by.keys())}")

    all_kw: list[str] = []
    seen: set[str] = set()
    for key in SHEET_KEYS:
        for p in rows_by[key]:
            k = p["keyword"]
            if k not in seen:
                seen.add(k)
                all_kw.append(k)

    month_label = resolve_month_label(xlsx)
    return {
        "monthLabel": month_label,
        "scoringYear": 2026,
        "blogRequirePostMonth": True,
        "hospitalName": "클래스병원",
        "hospitalNames": ["클래스병원"],
        "hospitalDomains": ["https://class2023.co.kr/"],
        "hospitalBlogBases": [
            "https://blog.naver.com/classlim2",
            "https://blog.naver.com/class231004",
        ],
        "sourceFileNote": str(xlsx.name),
        "manualRanks": {},
        "manualRanksByTab": {},
        "regionDefault": "",
        "keywords": all_kw,
        "rowsBySheetKey": rows_by,
        "sheetTitles": {
            "regional-pc": "2026 지역 PC",
            "regional-mob": "2026 지역 모바일",
            "national-pc": "2026 전국 PC",
            "national-mob": "2026 전국 모바일",
        },
    }


def main() -> None:
    xlsx = Path((os.getenv("CLASS_KEYWORD_XLSX") or DEFAULT_XLSX).strip())
    if not xlsx.exists():
        raise SystemExit(f"엑셀을 찾을 수 없습니다: {xlsx}")
    print("병합 슬롯: monthLabel =", resolve_month_label(xlsx), "| hospitalName = 클래스병원 (포인트병원 데이터와 별도)")
    cfg = build_config_from_xlsx(xlsx)
    print("키워드 수(중복 제거):", len(cfg["keywords"]))
    print("시트별 행 수:", {k: len(v) for k, v in cfg["rowsBySheetKey"].items()})
    bm = _load_build_module()
    bm.run_scoring_pipeline(cfg)
    print("완료: scoring-data.json 병합, last-run-evidence.json 의 byHospital['클래스병원'] 갱신")


if __name__ == "__main__":
    main()
