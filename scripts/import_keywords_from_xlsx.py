# -*- coding: utf-8 -*-
"""엑셀 배점표에서 지역/전국·PC/MOB 시트별 (지역, 키워드) 행을 추출해 채점용 JSON 설정을 만든다."""
from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]

def find_header_row(ws) -> int | None:
    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row or len(row) < 3:
            continue
        b = row[1]
        c = row[2]
        bs = str(b).strip() if b is not None else ""
        cs = str(c).strip() if c is not None else ""
        if bs == "지역" and cs == "키워드":
            return ri
    return None


def extract_pairs(ws) -> list[list[str | None]]:
    hdr = find_header_row(ws)
    if not hdr:
        return []
    out: list[list[str | None]] = []
    for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
        if not row or len(row) < 3:
            continue
        _no, reg, kw = row[0], row[1], row[2]
        kws = str(kw).strip() if kw is not None else ""
        if not kws or kws == "키워드" or "합계" in kws:
            continue
        reg_s = str(reg).strip() if reg is not None else ""
        if reg_s == "지역":
            reg_s = ""
        out.append([reg_s or None, kws])
    return out


def pairs_equal(a: list, b: list) -> bool:
    if len(a) != len(b):
        return False
    for x, y in zip(a, b):
        if (x[0] or "") != (y[0] or "") or x[1] != y[1]:
            return False
    return True


def dedupe_keywords_ordered(regional: list, national: list) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for _r, kw in regional + national:
        if kw in seen:
            continue
        seen.add(kw)
        out.append(kw)
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", type=Path, help="배점표 xlsx 경로")
    ap.add_argument(
        "-o",
        "--output",
        type=Path,
        default=ROOT / "config" / "march_from_xlsx_keywords.json",
        help="저장할 JSON (기본: config/march_from_xlsx_keywords.json)",
    )
    ap.add_argument("--month-label", default="3월")
    args = ap.parse_args()

    wb = load_workbook(args.xlsx, data_only=True)
    names = wb.sheetnames
    if len(names) < 4:
        raise SystemExit(f"시트가 4개 미만입니다: {names}")

    regional_pc = extract_pairs(wb[names[0]])
    regional_mob = extract_pairs(wb[names[1]])
    national_pc = extract_pairs(wb[names[2]])
    national_mob = extract_pairs(wb[names[3]])

    if not regional_pc:
        raise SystemExit("지역 PC 시트에서 데이터 행을 찾지 못했습니다.")
    if not pairs_equal(regional_pc, regional_mob):
        print("경고: 지역 PC/MOB 행이 다릅니다. 지역 PC 기준으로 MOB에도 동일 복사합니다.")
        regional_mob = [list(x) for x in regional_pc]
    if not national_pc:
        raise SystemExit("전국 PC 시트에서 데이터 행을 찾지 못했습니다.")
    if not pairs_equal(national_pc, national_mob):
        print("경고: 전국 PC/MOB 행이 다릅니다. 전국 PC 기준으로 MOB에도 동일 복사합니다.")
        national_mob = [list(x) for x in national_pc]

    rows_by_key = {
        "regional-pc": regional_pc,
        "regional-mob": regional_mob,
        "national-pc": national_pc,
        "national-mob": national_mob,
    }
    titles = {
        "regional-pc": str(names[0]).strip(),
        "regional-mob": str(names[1]).strip(),
        "national-pc": str(names[2]).strip(),
        "national-mob": str(names[3]).strip(),
    }

    keywords = dedupe_keywords_ordered(regional_pc, national_pc)

    cfg = {
        "monthLabel": args.month_label,
        "sourceFileNote": args.xlsx.name,
        "hospitalNames": ["포인트병원"],
        "manualRanks": {},
        "manualRanksByTab": {},
        "regionDefault": "",
        "keywords": keywords,
        "rowsBySheetKey": rows_by_key,
        "sheetTitles": titles,
        "scoringNote": "API 자동 탭(지도/카페/블로그/보도자료/웹)은 1위=3점, 2~5위=2점, 6~10위=1점, 미노출=0점.",
    }

    out_path = args.output
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
    print("저장:", out_path)
    print("지역 행:", len(regional_pc), "전국 행:", len(national_pc), "고유 키워드:", len(keywords))


if __name__ == "__main__":
    main()
