# -*- coding: utf-8 -*-
"""Export 배점표 xlsx to data/scoring-data.json for the static web app."""
import json
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
FILES = [
    ("1월", "포인트병원_1월_배점표.xlsx"),
    ("2월", "포인트병원_2월_배점표.xlsx"),
    ("3월", "포인트병원_3월_배점표.xlsx"),
]

HEADER_ROW = 3  # 0-based: row 4 in Excel
FIRST_DATA_ROW = 4


def cell_json(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, float) and v == int(v):
        return int(v)
    return v


def normalize_sheet_title(name: str) -> str:
    return name.strip()


def sheet_key(name: str) -> str:
    if "지역" in name and "PC" in name and "MOB" not in name:
        return "regional-pc"
    if "지역" in name and "MOB" in name:
        return "regional-mob"
    if "전국" in name and "PC" in name and "MOB" not in name:
        return "national-pc"
    if "전국" in name and "MOB" in name:
        return "national-mob"
    return re.sub(r"[^a-z0-9]+", "-", name.strip().lower()).strip("-") or "sheet"


def parse_workbook(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {"sourceFile": path.name, "sheets": []}
    for raw_name in wb.sheetnames:
        ws = wb[raw_name]
        title = normalize_sheet_title(raw_name)
        rows_iter = list(ws.iter_rows(values_only=True))
        if len(rows_iter) <= HEADER_ROW:
            continue
        header = [str(c).strip() if c is not None else "" for c in rows_iter[HEADER_ROW]]
        data_rows = []
        for row in rows_iter[FIRST_DATA_ROW:]:
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            data_rows.append([cell_json(c) for c in row])
        out["sheets"].append(
            {
                "key": sheet_key(raw_name),
                "title": title,
                "header": header,
                "rows": data_rows,
            }
        )
    return out


def main():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    months = []
    for label, fname in FILES:
        p = ROOT / fname
        if not p.exists():
            print("skip missing:", p)
            continue
        blob = parse_workbook(p)
        blob["monthLabel"] = label
        months.append(blob)
    payload = {"version": 1, "generatedBy": "export_xlsx_to_json.py", "months": months}
    out_path = DATA_DIR / "scoring-data.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print("Wrote", out_path, "months:", len(months))


if __name__ == "__main__":
    main()
