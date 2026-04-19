# -*- coding: utf-8 -*-
"""
두발로병원 키워드 엑셀: 지역·키워드 열만 읽고 네이버 API로 재채점 후 scoring-data.json 에 병합.

- 병합 슬롯은 (monthLabel, hospitalName) 이므로 다른 병원 데이터와 섞이지 않음.
- 엑셀 경로: DUBALO_KEYWORD_XLSX (미설정 시 기본 경로 사용)
- 월 라벨: DUBALO_MONTH_LABEL (예: 3월) 또는 파일명에서 N월 자동 추론
- 지역 규칙: '전국' 포함이면 전국, 나머지는 모두 '타지역'
"""
from __future__ import annotations

import importlib.util
import os
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SHEET_KEYS = ("regional-pc", "regional-mob", "national-pc", "national-mob")
DEFAULT_XLSX = r"y:\@컨텐츠팀\D_두발로병원\02. 월간보고서\202603\두발로병원_배점표_3월.xlsx"


def _load_build_module():
    path = ROOT / "scripts" / "build_april_month.py"
    spec = importlib.util.spec_from_file_location("build_april_month", path)
    if spec is None or spec.loader is None:
        raise RuntimeError("build_april_month 로드 실패")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def resolve_month_label(xlsx: Path) -> str:
    env = (os.getenv("DUBALO_MONTH_LABEL") or "").strip()
    if env:
        return env
    m = re.search(r"(\d{1,2})월", xlsx.name)
    if m:
        return f"{int(m.group(1))}월"
    return "3월"


def normalize_region(raw: str | None) -> str:
    t = (raw or "").strip()
    if "전국" in t:
        return "전국"
    return "타지역"


def sheet_key_from_excel_name(name: str) -> str | None:
    n = (name or "").replace(" ", "").strip()
    if not n:
        return None
    # 두발로 엑셀은 시트명에 '전국'이 함께 들어와도 실제 키워드는 지역 성격인 경우가 있어
    # 우선 디바이스 기준으로 regional 키에 적재하고, national은 아래에서 복제 생성한다.
    is_mob = ("모바일" in n) or ("MOB" in n.upper())
    return "regional-mob" if is_mob else "regional-pc"


def find_header_and_cols(ws) -> tuple[int, int, int] | None:
    max_r = min(ws.max_row or 0, 40)
    for r in range(1, max_r + 1):
        reg_c = kw_c = None
        for c in range(1, min(ws.max_column or 0, 40) + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).replace(" ", "").strip()
            if s == "지역":
                reg_c = c
            if "키워드" in s and "연관" not in s and "검색어" not in s and "합계" not in s:
                kw_c = c
        if reg_c and kw_c:
            return r, reg_c, kw_c
    return None


def read_pairs_keyword_only(ws) -> list[dict[str, str]]:
    found = find_header_and_cols(ws)
    if not found:
        return []
    header_r, reg_c, kw_c = found
    out: list[dict[str, str]] = []
    for r in range(header_r + 1, (ws.max_row or 0) + 1):
        raw_kw = ws.cell(r, kw_c).value
        if raw_kw is None or str(raw_kw).strip() == "":
            continue
        kw = str(raw_kw).strip()
        raw_reg = ws.cell(r, reg_c).value
        out.append({"region": normalize_region(str(raw_reg or "")), "keyword": kw})
    return out


def dedupe_pairs(pairs: list[dict[str, str]]) -> list[dict[str, str]]:
    out: list[dict[str, str]] = []
    seen: set[str] = set()
    for p in pairs:
        kw = (p.get("keyword") or "").strip()
        if not kw or kw in seen:
            continue
        seen.add(kw)
        out.append({"region": (p.get("region") or "타지역").strip() or "타지역", "keyword": kw})
    return out


def build_config_from_xlsx(xlsx: Path) -> dict[str, Any]:
    wb = load_workbook(xlsx, read_only=False, data_only=True)
    rows_by: dict[str, list[dict[str, str]]] = {k: [] for k in SHEET_KEYS}
    for raw_name in wb.sheetnames:
        key = sheet_key_from_excel_name(raw_name)
        if key not in SHEET_KEYS:
            continue
        ws = wb[raw_name]
        pairs = read_pairs_keyword_only(ws)
        if not pairs:
            continue
        rows_by[key].extend(pairs)
    wb.close()

    for key in SHEET_KEYS:
        rows_by[key] = dedupe_pairs(rows_by[key])

    # 전국 시트가 비어 있으면 동일 디바이스 지역 키워드를 전국으로 복제
    if not rows_by["national-pc"] and rows_by["regional-pc"]:
        rows_by["national-pc"] = [{"region": "전국", "keyword": p["keyword"]} for p in rows_by["regional-pc"]]
    if not rows_by["national-mob"] and rows_by["regional-mob"]:
        rows_by["national-mob"] = [{"region": "전국", "keyword": p["keyword"]} for p in rows_by["regional-mob"]]

    missing = [k for k in SHEET_KEYS if not rows_by.get(k)]
    if missing:
        raise SystemExit(f"필수 시트 데이터가 비었습니다: {missing}.")

    all_kw: list[str] = []
    seen: set[str] = set()
    for key in SHEET_KEYS:
        for p in rows_by[key]:
            kw = p["keyword"]
            if kw not in seen:
                seen.add(kw)
                all_kw.append(kw)

    month_label = resolve_month_label(xlsx)
    return {
        "monthLabel": month_label,
        "scoringYear": 2026,
        "blogRequirePostMonth": True,
        "hospitalName": "두발로병원",
        "hospitalNames": ["두발로병원"],
        "hospitalDomains": ["https://www.dubalo.com/"],
        "hospitalBlogBases": [
            "https://blog.naver.com/dubalo2021",
            "https://blog.naver.com/dubaloklee",
            "https://blog.naver.com/seoulfootandankle",
            "https://blog.naver.com/compare11529",
        ],
        "sourceFileNote": str(xlsx.name),
        "manualRanks": {},
        "manualRanksByTab": {},
        "regionDefault": "타지역",
        "keywords": all_kw,
        "rowsBySheetKey": rows_by,
        "sheetTitles": {
            "regional-pc": "2026 지역 PC",
            "regional-mob": "2026 지역 모바일",
            "national-pc": "2026 전국 PC",
            "national-mob": "2026 전국 모바일",
        },
    }


def main() -> None:
    xlsx = Path((os.getenv("DUBALO_KEYWORD_XLSX") or DEFAULT_XLSX).strip())
    if not xlsx.exists():
        raise SystemExit(f"엑셀을 찾을 수 없습니다: {xlsx}")
    print("병합 슬롯: monthLabel =", resolve_month_label(xlsx), "| hospitalName = 두발로병원")
    cfg = build_config_from_xlsx(xlsx)
    print("키워드 수(중복 제거):", len(cfg["keywords"]))
    print("시트별 행 수:", {k: len(v) for k, v in cfg["rowsBySheetKey"].items()})
    bm = _load_build_module()
    bm.run_scoring_pipeline(cfg)
    print("완료: scoring-data.json 병합, last-run-evidence.json 의 byHospital['두발로병원'] 갱신")


if __name__ == "__main__":
    main()
